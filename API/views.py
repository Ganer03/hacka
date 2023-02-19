import json
import os
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from API.models import Landlord, Leaser, Building, CommercialObject, Position, Personal, Service, \
    HcsType, Hcs
from rest_framework import generics, viewsets, mixins
from API.serializers import LandLordSerializer, LeaserSerializer, BuildingSerializer, CommercialObjectSerializer, \
    PositionSerializer, PersonalSerializer, ServiceSerializer, HcsTypeSerializer, HcsSerializer
from openpyxl import Workbook
from openpyxl.styles import Font
from random import choices
from string import ascii_letters, digits

a = 0


def get_free_name():
    filename = "".join(choices(ascii_letters + digits, k=12)) + '.xlsx'
    path = os.path.join('reports', filename)
    if os.path.isfile(path):
        return get_free_name()
    return path


def counter_agents(request):
    ft = Font(bold=True)
    wb = Workbook()
    ws = wb.active
    ws.append(['Контрагент', 'ФИО', 'Адрес', 'Email', 'Телефон'])
    for row in ws["A1:Z1"]:
        for cell in row:
            cell.font = ft

    for leaser in Leaser.objects.all():
        ws.append([leaser.company_name, str(leaser), leaser.address, leaser.email, leaser.phone])

    path = get_free_name()
    wb.save(path)

    return JsonResponse({
        'path': f"http://{request.get_host()}/{path}"
    })


def counter_agents_json(request):
    res = []

    for leaser in Leaser.objects.all():
        res.append({
            'agent': leaser.company_name,
            'name': str(leaser),
            'address': leaser.address,
            'email': leaser.email,
            'phone': leaser.phone
        })

    return JsonResponse(res, safe=False)


def rent_documents(request, lord_pk):
    ft = Font(bold=True)
    wb = Workbook()
    ws = wb.active
    ws.append(['Контрагент', 'Телефон', 'Арендное соглашение'])
    for row in ws["A1:Z1"]:
        for cell in row:
            cell.font = ft

    landlord = get_object_or_404(Landlord.objects, pk=lord_pk)
    for building in Building.objects.filter(landlord=landlord):
        for obj in CommercialObject.objects.filter(building=building):
            if obj.leaser and obj.document:
                ws.append([obj.leaser.company_name, obj.leaser.phone,
                           f"http://{request.get_host()}/{obj.document}"])

    path = get_free_name()
    wb.save(path)

    return JsonResponse({
        'path': f"http://{request.get_host()}/{path}"
    })


def rent_documents_json(request, lord_pk):
    ws = []

    landlord = get_object_or_404(Landlord.objects, pk=lord_pk)
    for building in Building.objects.filter(landlord=landlord):
        for obj in CommercialObject.objects.filter(building=building):
            if obj.leaser and obj.document:
                ws.append({
                    "agent": obj.leaser.company_name,
                    'phone': obj.leaser.phone,
                    'document': f"http://{request.get_host()}/{obj.document}"
                })

    return JsonResponse(ws, safe=False)


class Landlords(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
                mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Landlord.objects.all()
    serializer_class = LandLordSerializer


class Leasers(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
              mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Leaser.objects.all()
    serializer_class = LeaserSerializer


class Buildings(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
                mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Building.objects.all()
    serializer_class = BuildingSerializer


class CommercialObjects(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin,
                        mixins.UpdateModelMixin,
                        mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = CommercialObject.objects.select_related('leaser')
    serializer_class = CommercialObjectSerializer


class Positions(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
                mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer


class Personals(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
                mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Personal.objects.all()
    serializer_class = PersonalSerializer


class Services(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
               mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer


class HcsTypes(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
               mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = HcsType.objects.all()
    serializer_class = HcsTypeSerializer


class Hcss(viewsets.GenericViewSet, mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
           mixins.DestroyModelMixin, mixins.CreateModelMixin):
    queryset = Hcs.objects.prefetch_related("hcs_type")
    serializer_class = HcsSerializer


# class Reviews(generics.ListAPIView):
#     queryset = Review.objects.select_related('author', 'doctor')
#     serializer_class = ReviewSerializer


def get_buildings(request, lord_pk):
    landlord = get_object_or_404(Landlord.objects, pk=lord_pk)
    return JsonResponse(list(Building.objects.filter(landlord=landlord).values()), safe=False,
                        json_dumps_params={'ensure_ascii': False})


def get_objects(request, building_pk):
    building = get_object_or_404(Building.objects, pk=building_pk)
    return JsonResponse(list(CommercialObject.objects.filter(building=building).values()), safe=False,
                        json_dumps_params={'ensure_ascii': False})


def get_board(request, lord_pk):
    global expence, incomer
    lalord = get_object_or_404(Landlord.objects, pk=lord_pk)
    hgs_water = [0 for x in range(0, 12)]
    hgs_heating = [0 for x in range(0, 12)]
    hgs_energy = [0 for x in range(0, 12)]
    for x in Building.objects.filter(landlord=lalord):
        rezult2 = CommercialObject.objects.filter(building=x).all()
        for y in rezult2:
            i, k, j = 0, 0, 0
            for item in Hcs.objects.filter(author=y.leaser):
                if item.hcs_type.name == 'Вода':
                    hgs_water[i] += item.value
                    i += 1
                elif item.hcs_type.name == 'Электричество':
                    hgs_energy[i] += item.value
                    j += 1
                elif item.hcs_type.name == 'Отопление':
                    hgs_heating[k] += item.value
                    k += 1
    return JsonResponse({
        'time': f'{hgs_water}',
        'energy': f'{hgs_energy}',
        'heating': f'{hgs_heating}'
    }, json_dumps_params={'ensure_ascii': False})


def get_incomes(request, lord_pk):
    landlord = Landlord.objects.filter(pk=lord_pk).all()
    incomer = 0
    expence = 0
    for x in landlord:
        incomer = x.incomer()
        expence = x.expense()
    return JsonResponse({
        'expence': f'{expence}',
        'incomer': f'{incomer}',
    }, json_dumps_params={'ensure_ascii': False})


def get_peoples(request):
    with open("a.json", 'r') as f:
        data = json.loads(f.read())
    return JsonResponse(data)


def set_peoples(request, a):
    data = {"count": a}
    with open("a.json", 'w') as f:
        f.write(json.dumps(data))
    return JsonResponse(data)
